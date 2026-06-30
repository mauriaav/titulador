"""
titulo_manager.py
-----------------
Responsabilidad: cargar títulos desde CSV/Excel y proveer búsqueda.
"""

import csv
import unicodedata
from dataclasses import dataclass
from pathlib import Path


def normalizar(texto: str) -> str:
    """Convierte a minúsculas y elimina tildes para comparación."""
    texto = texto.lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto


@dataclass
class Titulo:
    """
    Representa una entrada del CSV/Excel.

    Atributos
    ---------
    titulo : str
        Texto corto que se muestra y se usa para buscar.
    cuerpo : str
        Texto completo que se copia al portapapeles al seleccionar.
        Si no hay columna 'cuerpo' en el archivo, queda igual a `titulo`.
    """
    titulo: str
    cuerpo: str


class TituloManager:
    """
    Gestiona la colección de títulos: carga desde archivo y búsqueda fuzzy.

    Atributos
    ---------
    ruta_archivo : Path
        Ruta al CSV o Excel con los títulos.
    titulos : list[Titulo]
        Lista de entradas cargadas (título + cuerpo a copiar).
    """

    EXTENSIONES_VALIDAS = {".csv", ".xlsx", ".xls"}

    def __init__(self, ruta_archivo: str | Path):
        self.ruta_archivo = Path(ruta_archivo)
        self.titulos: list[Titulo] = []
        self.cargar()

    # ------------------------------------------------------------------
    # Carga
    # ------------------------------------------------------------------

    def cargar(self) -> None:
        """Carga los títulos desde el archivo. Detecta formato por extensión."""
        if not self.ruta_archivo.exists():
            raise FileNotFoundError(f"No se encontró el archivo: {self.ruta_archivo}")

        ext = self.ruta_archivo.suffix.lower()

        if ext not in self.EXTENSIONES_VALIDAS:
            raise ValueError(f"Formato no soportado: {ext}")

        if ext == ".csv":
            self._cargar_csv()
        else:
            self._cargar_excel()

    def _cargar_csv(self) -> None:
        # Prueba encodings en orden: UTF-8 con BOM, UTF-8, Windows-1252 (Excel español)
        for encoding in ("utf-8-sig", "utf-8", "windows-1252", "latin-1"):
            try:
                with open(self.ruta_archivo, encoding=encoding, newline="") as f:
                    reader = csv.DictReader(f)
                    fieldnames = reader.fieldnames or []
                    if "titulo" not in fieldnames:
                        raise ValueError("El CSV debe tener una columna llamada 'titulo'.")

                    tiene_cuerpo = "cuerpo" in fieldnames

                    self.titulos = []
                    for fila in reader:
                        titulo = fila["titulo"].strip()
                        if not titulo:
                            continue
                        cuerpo = fila["cuerpo"].strip() if tiene_cuerpo else titulo
                        self.titulos.append(Titulo(titulo=titulo, cuerpo=cuerpo or titulo))
                return  # Si llegó acá, anduvo bien
            except UnicodeDecodeError:
                continue
        raise ValueError("No se pudo leer el archivo CSV: encoding desconocido.")

    def _cargar_excel(self) -> None:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Instalá openpyxl para leer archivos Excel: pip install openpyxl")

        wb = openpyxl.load_workbook(self.ruta_archivo, read_only=True, data_only=True)
        ws = wb.active

        encabezados = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows())]
        if "titulo" not in encabezados:
            raise ValueError("El Excel debe tener una columna llamada 'titulo'.")

        col_titulo = encabezados.index("titulo")
        col_cuerpo = encabezados.index("cuerpo") if "cuerpo" in encabezados else None

        self.titulos = []
        for fila in ws.iter_rows(min_row=2, values_only=True):
            valor_titulo = fila[col_titulo]
            if not valor_titulo:
                continue
            titulo = str(valor_titulo).strip()

            if col_cuerpo is not None and fila[col_cuerpo]:
                cuerpo = str(fila[col_cuerpo]).strip()
            else:
                cuerpo = titulo

            self.titulos.append(Titulo(titulo=titulo, cuerpo=cuerpo))
        wb.close()

    def recargar(self) -> None:
        """Recarga los títulos desde el archivo (útil si fue editado)."""
        self.titulos.clear()
        self.cargar()

    # ------------------------------------------------------------------
    # Búsqueda
    # ------------------------------------------------------------------

    def buscar(self, query: str, max_resultados: int = 50) -> list[Titulo]:
        """
        Búsqueda tolerante a tildes y mayúsculas. Busca solo en el campo `titulo`.
        Prioriza: coincidencia al inicio > coincidencia en cualquier parte.

        Parámetros
        ----------
        query : str
            Texto ingresado por el usuario.
        max_resultados : int
            Cantidad máxima de resultados a devolver.

        Retorna
        -------
        list[Titulo]
            Entradas que coinciden, ordenadas por relevancia.
        """
        if not query.strip():
            return self.titulos[:max_resultados]

        query_norm = normalizar(query)
        tokens = query_norm.split()

        inicio = []
        medio = []

        for entrada in self.titulos:
            titulo_norm = normalizar(entrada.titulo)

            # Todos los tokens deben estar presentes
            if not all(t in titulo_norm for t in tokens):
                continue

            if titulo_norm.startswith(tokens[0]):
                inicio.append(entrada)
            else:
                medio.append(entrada)

        return (inicio + medio)[:max_resultados]

    def __len__(self) -> int:
        return len(self.titulos)

    def __repr__(self) -> str:
        return f"TituloManager(archivo='{self.ruta_archivo}', títulos={len(self)})"
